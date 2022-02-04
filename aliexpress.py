from calendar import c
from textwrap import indent
import requests
from openpyxl import Workbook, load_workbook
import os
from bs4 import BeautifulSoup
import matplotlib.pyplot as plt

def main():
    execelFileName = "product.xlsx"
    page = "3"
    global excel, sheet
    # excel eke  write krnn
    if os.path.isfile(execelFileName):
        excel = load_workbook(filename=execelFileName)
        print("file loaded !")
        sheet = excel.active
        sheetName = "page " + page
        sheet = excel.create_sheet(title=sheetName)
        # sheet.title = "page " + page

    else:
        excel = Workbook()
        sheet = excel.active
        sheet.title = "page " + page
        sheet.append(["ID", " ", "Product Details", " ", "Date"])
        print("new file created !")

    try:
        try:
            dm = "https://www.awwwards.com/"
            prm = "awwwards/collections/product-page/?page="
            url = dm + prm + page
            response = requests.get(url)
            response.raise_for_status()
            # get all the html page
            # soup = BeautifulSoup(response.content, "html.parser")
            # body = soup.body
            soup = BeautifulSoup(response.text, "html.parser")
            # print(soup)
            bodyList = soup.find("ul", class_="list-items").find_all(
                "li", class_="col-3 js-collectable"
            )
            # Print each string recursively
            print(len(bodyList))
            # print(type(body))
            # for item in bodyList:
            for index, item in enumerate(bodyList):
                name = item.find("div", class_="box-item").h3.text
                date = (
                    item.find("div", class_="box-item")
                    .find("div", class_="box-right")
                    .text
                )
                # print(name + "--" + date)
                if name != " " and date != " ":
                    sheet.append([index, " ", name, " ", date])
            # end of the loop save file
            excel.save(filename=execelFileName)
        except Exception as e:
            print(e)
    except Exception as e:
        print(e)


if __name__ == "__main__":
    main()
